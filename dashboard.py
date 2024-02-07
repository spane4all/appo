import streamlit as st
import pandas as pd
from io import BytesIO
from PyPDF2 import PdfReader
import io
from openpyxl import load_workbook
import re

# Global variable to store backup history
backup_history = []

def extract_information(text):
    atc_region = re.search(r'ATC REGION\s+(.+)', text).group(1).strip()
    lessee_site_number = re.search(r'LESSEE  SITE NUMBER\s+(.+)', text).group(1).strip()
    total_per_month = re.search(r'TOTAL  -  per month \(exclusive of VAT\) (.+?) R', text).group(1).strip()
    renewal_term_commencement_date = re.search(r'RENEWAL TERM  COMMENCEMENT  DATE (.+)', text).group(1).strip()

    return atc_region, lessee_site_number, total_per_month, renewal_term_commencement_date

def process_pdf(pdf_file):
    pdf_data = []
    with pdf_file as file:
        pdf_reader = PdfReader(file)
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            pdf_data.append(page.extract_text())
    return pdf_data

def process_excel(excel_file, pdf_data):
    global backup_history

    try:
        # Load Excel workbook using openpyxl
        workbook = load_workbook(excel_file)

        # Select the "Additions and Modification" sheet if it exists, otherwise create it
        sheet_name = "Additions and Modification"
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)

        sheet = workbook[sheet_name]

        # Extract specific information from PDF
        atc_region, lessee_site_number, total_per_month, renewal_term_commencement_date = extract_information(pdf_data[0])

        # Identify the next empty columns for each field
        headers = [cell.value for cell in sheet[1]]
        region_column = headers.index('Region') + 1 if 'Region' in headers else len(headers) + 1
        external_asset_id_column = headers.index('External Asset ID') + 1 if 'External Asset ID' in headers else len(headers) + 2
        minimum_lease_payment_column = headers.index('MinimumLeasePaymentaspercontract') + 1 if 'MinimumLeasePaymentaspercontract' in headers else len(headers) + 3
        current_lease_commencement_column = headers.index('Current Lease Commencement Date') + 1 if 'Current Lease Commencement Date' in headers else len(headers) + 4

        # Find the next empty row
        next_row = sheet.max_row + 1

        # Update Excel sheet with extracted information in the next empty columns
        sheet.cell(row=next_row, column=region_column, value=atc_region)
        sheet.cell(row=next_row, column=external_asset_id_column, value=lessee_site_number)
        sheet.cell(row=next_row, column=minimum_lease_payment_column, value=total_per_month)
        sheet.cell(row=next_row, column=current_lease_commencement_column, value=renewal_term_commencement_date)

        # Save changes to Excel file
        buffer = BytesIO()
        workbook.save(buffer)

        # Print to check if backup_history is being updated
        print("Backup history before update:", backup_history)

        # Update backup history
        backup_history.append({
            'ATC REGION': atc_region,
            'LESSEE SITE NUMBER': lessee_site_number,
            'TOTAL - per month (exclusive of VAT)': total_per_month,
            'RENEWAL TERM COMMENCEMENT DATE': renewal_term_commencement_date,
            'File Name': excel_file.name  # Add the file name to the history
        })

        # Print to check if backup_history is updated
        print("Backup history after update:", backup_history)

        # Create DataFrame for display
        df = pd.DataFrame({'ATC REGION': [atc_region],
                           'LESSEE SITE NUMBER': [lessee_site_number],
                           'TOTAL - per month (exclusive of VAT)': [total_per_month],
                           'RENEWAL TERM COMMENCEMENT DATE': [renewal_term_commencement_date]})

        return df, buffer

    except Exception as e:
        print(f"Error in process_excel: {e}")
        return pd.DataFrame(), None



def pdf():
    global backup_history

    st.title("PDF to Excel Streamlit App")

    # File Upload for PDF
    st.sidebar.header("Upload PDF")
    pdf_file = st.sidebar.file_uploader("Choose a PDF file", type=["pdf"])

    # File Upload for Excel
    st.sidebar.header("Upload Excel Template")
    excel_file = st.sidebar.file_uploader("Choose an Excel file", type=["xlsx"])

    if pdf_file and excel_file:
        # Process PDF
        pdf_data = process_pdf(pdf_file)

        # Process Excel
        buffer = process_excel(excel_file, pdf_data)

        if buffer:
            # Display PDF data
            st.header("PDF Data")
            st.write(pdf_data)

            # Download button for the duplicated file
            download_btn = st.button("Download Duplicated Excel File")
            if download_btn:
                st.info("Processing... Please wait.")
                st.download_button(
                    label="Click here to download",
                    data=buffer.getvalue(),
                    file_name=f"duplicated_{excel_file.name}",  # Keep the same name as the original file
                    key="download_duplicated_file"
                )

            # Update backup history page
            backup_history_df = pd.DataFrame(backup_history)
            st.header("Backup History")
            st.write(backup_history_df)

        else:
            st.error("Error processing Excel. Please check the console for details.")


def backup_page():
    global backup_history  # Add this line to explicitly use the global variable
    st.title("Backup History Page")

    if not backup_history:
        st.info("No backup history available.")
    else:
        # Display backup history in a table
        history_df = pd.DataFrame(backup_history)
        st.write(history_df)


def about_page():
    st.title("About the App")

    about_text = [
        "This app is designed to help teams manage lease agreements. It provides features such as:",
        "- Uploading lease agreements",
        "- Extracting information from PDFs",
        "- Recording data in an input sheet",
        "- Data backup functionality",
        "- Notification features for important dates",
        "Feel free to explore the different pages and features!"
    ]

    for line in about_text:
        st.markdown(line)

def main():
    global excel_file  # Ensure we use the global excel_file variable

    st.set_page_config(page_title="Lease Agreement Management App")

    pages = {
        "About the App": about_page,
        "PDF to Excel": pdf,
        "Backup History": backup_page  # Added backup history page
    }

    st.sidebar.title("Navigation")
    selected_page = st.sidebar.radio("Go to", list(pages.keys()))

    # Run the selected page function
    pages[selected_page]()

if __name__ == "__main__":
    main()
